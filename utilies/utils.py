
import csv
import inspect
import logging
import softest
from openpyxl import Workbook, load_workbook


class utils(softest.TestCase):
    def assertlistitemtext(self,list,value):

        for stop in list:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEquals, stop.text, value)
            if stop.text == value:
                print("test pass")
            else:
                print("test failed")

        self.assert_all()

    def Custom_logger(logLevel=logging.DEBUG):
        # set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        # create console handler or file handler and set the log level
        # console_handler = logging.StreamHandler()
        file_handler = logging.FileHandler("automation.log")

        # create formatter - how you want your logs to be formatted
        formater = logging.Formatter('%(asctime)s - %(levelname)s : %(message)s',
                                     datefmt='%m-%d-%Y %I:%M:%S %p')

        # add formatter to console or file handler
        # console_handler.setFormatter(formater)
        file_handler.setFormatter(formater)

        # add console handler to logger
        # logger.addHandler(console_handler)
        logger.addHandler(file_handler)
        return logger



    def read_data_xlsx(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]

        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist


    def read_data_from_csv(filename):
        #create a empty list
        datalist= []

        # open csv file
        csvdata = open(filename,"r")

        # create a csv reader
        reader = csv.reader(csvdata)

        # skip header
        next(reader)

        # add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist
